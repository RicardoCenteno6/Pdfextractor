from flask import Flask, render_template, request, send_file
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
from datetime import datetime

app = Flask(__name__)

# Configuración de carpetas
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def extraer_datos_pdf(pdf_path):
    """
    Extrae información específica del PDF según los campos definidos.
    Filtra las instrucciones y captura solo las respuestas del usuario.
    
    Args:
        pdf_path: Ruta del archivo PDF a procesar
        
    Returns:
        dict: Diccionario con los campos extraídos
    """
    # Inicializamos el diccionario con los campos vacíos
    datos = {
        'Nombre del solicitante': '',
        'Número del expediente': '',
        'Correo electrónico': '',
        'Número de Resolución a revisar': '',
        'Nueva prueba o argumento': '',
        'Motivo de la revisión': ''
    }
    
    # Frases instructivas que debemos ignorar (no son respuestas del usuario)
    frases_a_ignorar = [
        'Detalle la nueva prueba o argumento',
        'Describa de forma breve el motivo',
        'Escriba el número de la Resolución',
    ]
    
    try:
        # Abrimos el PDF con PyMuPDF
        doc = fitz.open(pdf_path)
        texto_completo = ""
        
        # Extraemos el texto de todas las páginas
        for pagina in doc:
            texto_completo += pagina.get_text()
        
        # Cerramos el documento
        doc.close()
        
        # DEBUG: Imprimir el texto completo (descomenta para ver qué captura)
        # print("="*50)
        # print("TEXTO EXTRAÍDO DEL PDF:")
        # print(texto_completo)
        # print("="*50)
        
        # NUEVA ESTRATEGIA:
        # 1. Buscamos el campo (encabezado)
        # 2. Capturamos la línea inmediatamente después
        # 3. Verificamos que NO sea texto instructivo
        # 4. Si la primera línea es instructiva, buscamos la siguiente línea con contenido
        
        # 1. Nombre del solicitante
        patron = r'Nombre del solicitante\s*\n\s*(.+?)(?=\n|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE | re.DOTALL)
        if match:
            respuesta = match.group(1).strip()
            # Limpiamos saltos de línea excesivos
            respuesta = re.sub(r'\n+', ' ', respuesta)
            datos['Nombre del solicitante'] = respuesta
        
        # 2. Número del expediente
        patron = r'Número del expediente\s*\n\s*(.+?)(?=\n|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE)
        if match:
            respuesta = match.group(1).strip()
            datos['Número del expediente'] = respuesta
        
        # 3. Correo electrónico
        patron = r'Correo electrónico\s*\n\s*(.+?)(?=\n|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE)
        if match:
            respuesta = match.group(1).strip()
            datos['Correo electrónico'] = respuesta
        
        # 4. Número de Resolución a revisar
        patron = r'Número de Resolución a revisar\s*\n(.+?)(?=Escriba el número|Nueva prueba|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE | re.DOTALL)
        if match:
            texto_capturado = match.group(1).strip()
            # Dividimos en líneas y filtramos las instrucciones
            lineas = texto_capturado.split('\n')
            lineas_filtradas = []
            for linea in lineas:
                linea = linea.strip()
                # Ignoramos líneas vacías y líneas con instrucciones
                if linea and not any(frase.lower() in linea.lower() for frase in frases_a_ignorar):
                    lineas_filtradas.append(linea)
            # Tomamos solo la primera línea válida (la respuesta)
            if lineas_filtradas:
                datos['Número de Resolución a revisar'] = lineas_filtradas[0]
        
        # 5. Nueva prueba o argumento
        # Buscamos después del encabezado, saltando la línea de instrucciones
        patron = r'Nueva prueba o argumento\s*\n(.+?)(?=Detalle la nueva prueba|Motivo de la revisión|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE | re.DOTALL)
        if match:
            texto_capturado = match.group(1).strip()
            # Dividimos en líneas y filtramos las instrucciones
            lineas = texto_capturado.split('\n')
            lineas_filtradas = []
            for linea in lineas:
                linea = linea.strip()
                # Ignoramos líneas vacías y líneas con instrucciones
                if linea and not any(frase.lower() in linea.lower() for frase in frases_a_ignorar):
                    lineas_filtradas.append(linea)
            datos['Nueva prueba o argumento'] = ' '.join(lineas_filtradas)
        
        # 6. Motivo de la revisión
        patron = r'Motivo de la revisión\s*\n(.+?)(?=Describa de forma breve|Recuerde que|$)'
        match = re.search(patron, texto_completo, re.IGNORECASE | re.DOTALL)
        if match:
            texto_capturado = match.group(1).strip()
            # Dividimos en líneas y filtramos las instrucciones
            lineas = texto_capturado.split('\n')
            lineas_filtradas = []
            for linea in lineas:
                linea = linea.strip()
                # Ignoramos líneas vacías y líneas con instrucciones
                if linea and not any(frase.lower() in linea.lower() for frase in frases_a_ignorar):
                    lineas_filtradas.append(linea)
            datos['Motivo de la revisión'] = ' '.join(lineas_filtradas)
        
    except Exception as e:
        print(f"Error procesando {pdf_path}: {str(e)}")
    
    return datos

def crear_excel(lista_datos, output_path):
    """
    Crea un archivo Excel con los datos extraídos de múltiples PDFs.
    
    Args:
        lista_datos: Lista de diccionarios con los datos extraídos
        output_path: Ruta donde se guardará el Excel
    """
    # Creamos un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Extraídos"
    
    # Definimos los encabezados (nombres de columnas)
    encabezados = [
        'Nombre del solicitante',
        'Número del expediente',
        'Correo electrónico',
        'Número de Resolución a revisar',
        'Nueva prueba o argumento',
        'Motivo de la revisión'
    ]
    
    # Escribimos los encabezados en la primera fila
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        # Aplicamos formato: negrita y fondo azul claro
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribimos los datos de cada PDF en las filas siguientes
    for fila_idx, datos in enumerate(lista_datos, start=2):
        for col_idx, campo in enumerate(encabezados, start=1):
            valor = datos.get(campo, '')
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Ajustamos el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardamos el archivo Excel
    wb.save(output_path)

@app.route('/')
def index():
    """Página principal donde se suben los PDFs"""
    return render_template('index.html')

@app.route('/procesar', methods=['POST'])
def procesar():
    """
    Procesa los archivos PDF subidos y genera el Excel.
    Esta función se ejecuta cuando el usuario hace clic en "Procesar PDFs"
    """
    # Verificamos si se subieron archivos
    if 'archivos' not in request.files:
        return "No se subieron archivos", 400
    
    archivos = request.files.getlist('archivos')
    
    if not archivos or archivos[0].filename == '':
        return "No se seleccionaron archivos", 400
    
    lista_datos = []
    archivos_procesados = []
    
    # Procesamos cada archivo PDF
    for archivo in archivos:
        if archivo and archivo.filename.endswith('.pdf'):
            # Guardamos el archivo temporalmente
            filename = archivo.filename
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            archivo.save(filepath)
            
            # Extraemos los datos del PDF
            datos = extraer_datos_pdf(filepath)
            lista_datos.append(datos)
            archivos_procesados.append(filename)
            
            # Eliminamos el archivo temporal
            os.remove(filepath)
    
    if not lista_datos:
        return "No se procesaron archivos PDF válidos", 400
    
    # Creamos el archivo Excel con los datos extraídos
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"datos_extraidos_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    
    crear_excel(lista_datos, output_path)
    
    # Enviamos el archivo Excel al usuario para descarga
    return send_file(
        output_path,
        as_attachment=True,
        download_name=output_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    # Iniciamos el servidor en modo desarrollo
    # debug=True permite ver errores detallados
    app.run(debug=True, port=5000)